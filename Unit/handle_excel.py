#coding=utf-8
import os
import sys
import openpyxl
base_path = os.path.abspath(os.path.dirname(__file__)).split('Unit')[0]
sys.path.append(base_path)

class HandleExcel:
    def __init__(self,file_name=None):
        if file_name:
            self.file_name = file_name
        else:
            self.file_name = base_path + "Case/aglione_interface.xlsx"

    def     load_excel(self):
        open_excel = openpyxl.load_workbook(self.file_name)
        return open_excel

    def get_sheet_data(self,index=None):
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 0
        data =  self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,cols):
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data

    def get_rows(self):
        rows = self.get_sheet_data().max_row
        return rows

    def get_row_data(self,row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self,row,cols,value):
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(base_path + "/Case/aglione_interface.xlsx")

    def get_columns_value(self,key=None):
        columns_list = []
        if key == None:
            key = "A"
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self,case_id):
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num += 1
        return num

    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_row_data(i+2))
        return data_list

hd_excel = HandleExcel()
if __name__ == '__main__':
    base_path = os.path.abspath(os.path.dirname(__file__)).split('Unit')[0]
    sys.path.append(base_path)
    excel_path = base_path + "Case/aglione_interface.xlsx"
    print(excel_path)

    excel = HandleExcel()
    x = excel.load_excel()
    print(x)

